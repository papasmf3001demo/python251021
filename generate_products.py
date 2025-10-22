import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "제품목록"

# 헤더
headers = ["제품ID", "제품명", "가격", "수량"]
ws.append(headers)

product_types = ["TV", "냉장고", "세탁기", "에어컨", "노트북", "스마트폰", "태블릿", "헤드폰", "스피커", "카메라"]

# 100개 생성
for i in range(1, 101):
    #서식문자를 지정 
    pid = f"P{i:04d}"
    name = f"{random.choice(product_types)} {random.randint(100, 999)}"
    price = random.randint(10_000, 1_500_000)
    qty = random.randint(1, 200)
    ws.append([pid, name, price, qty])

# 열 너비 조정
for col in range(1, 5):
    ws.column_dimensions[get_column_letter(col)].width = 18

save_path = r"c:\work\ProductList.xlsx"
wb.save(save_path)
print(f"Saved: {save_path}")